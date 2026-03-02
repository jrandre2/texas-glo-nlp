#!/usr/bin/env python3
"""
Independent lineage audit from original XLSX sources to activity-level output.

This validator reads raw source workbooks directly (without using parser helpers),
rebuilds the P31/F31 wide structures, and compares them against:
  - P31_Restructured
  - F31_Restructured
  - Merged_Activity_1to1

Outputs:
  - output/spreadsheet/Activity_Lineage_Audit.csv
  - output/spreadsheet/Activity_Lineage_Audit.xlsx
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "output" / "spreadsheet" / "Activity_Level_Analytic_Dataset.xlsx"
DEFAULT_AUDIT_CSV = ROOT / "output" / "spreadsheet" / "Activity_Lineage_Audit.csv"
DEFAULT_AUDIT_XLSX = ROOT / "output" / "spreadsheet" / "Activity_Lineage_Audit.xlsx"
YEARS = list(range(2020, 2027))

SOURCE_PREFIXES = {
    "B17_F31": "B-17-DM-48-0001_F31",
    "B18_F31": "B-18-DP-48-0001_F31",
    "HIM1_P31": "P-17-TX-48-HIM1_P31",
}


@dataclass
class Check:
    check: str
    status: str
    expected: Any
    actual: Any
    delta: float
    notes: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def clean_num(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value == value else None
    s = clean_text(value)
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace(",", "").replace("$", "").strip()
    if s in {"", "-", "--"}:
        return None
    try:
        out = float(s)
        return -out if neg else out
    except ValueError:
        return None


def sanitize_token(value: Any) -> str:
    s = clean_text(value)
    if not s:
        return "Unknown"
    s = s.replace("&", "and")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "Unknown"


def year_from_quarter_label(value: Any) -> Optional[int]:
    s = clean_text(value)
    m = re.match(r"^(\d{4})\s*Q([1-4])$", s)
    if m:
        return int(m.group(1))
    if isinstance(value, datetime):
        return int(value.year)
    return None


def preferred_candidate(paths: List[Path]) -> Path:
    def rank(p: Path) -> Tuple[int, int, str]:
        has_copy = 1 if re.search(r"\[\d+\](?=\.xlsx$)", p.name) else 0
        return (has_copy, len(p.name), p.name)

    return sorted(paths, key=rank)[0]


def resolve_sources(root: Path) -> Dict[str, Path]:
    xlsx = [p for p in root.glob("*.xlsx") if p.is_file() and not p.name.startswith("._")]
    resolved: Dict[str, Path] = {}
    for key, prefix in SOURCE_PREFIXES.items():
        matches = [p for p in xlsx if p.name.startswith(prefix)]
        if not matches:
            raise FileNotFoundError(f"Missing source workbook for {key} with prefix {prefix}")
        resolved[key] = preferred_candidate(matches)
    return resolved


def best_org_map(df: pd.DataFrame, activity_col: str, org_col: str) -> pd.DataFrame:
    tmp = df[[activity_col, org_col]].copy()
    tmp[activity_col] = tmp[activity_col].astype(str).str.strip()
    tmp[org_col] = tmp[org_col].fillna("").astype(str).str.strip()
    tmp = tmp[tmp[activity_col] != ""]
    rows = []
    for activity, g in tmp.groupby(activity_col, dropna=False):
        vc = g[g[org_col] != ""][org_col].value_counts()
        org = vc.index[0] if not vc.empty else ""
        rows.append({"Activity Number": activity, "Activity Responsible Org": org})
    return pd.DataFrame(rows)


def read_p31_raw(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    qlabel_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    quarter_cols: List[Tuple[int, Optional[str], Optional[int]]] = []
    for c in range(6, len(qlabel_row)):
        qlabel = clean_text(qlabel_row[c])
        year = year_from_quarter_label(qlabel)
        if year in YEARS:
            quarter_cols.append((c, qlabel, year))

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        activity_number = clean_text(row[2] if len(row) > 2 else None)
        if not activity_number:
            continue
        activity_type = clean_text(row[0] if len(row) > 0 else None)
        org = clean_text(row[1] if len(row) > 1 else None)
        measure_type = clean_text(row[4] if len(row) > 4 else None)
        metric = clean_text(row[5] if len(row) > 5 else None)

        for c, qlabel, year in quarter_cols:
            if c >= len(row):
                continue
            v = clean_num(row[c])
            if v is None:
                continue
            rows.append(
                {
                    "Activity Number": activity_number,
                    "Activity Type": activity_type,
                    "Activity Measure Type": measure_type,
                    "Metrics": metric,
                    "Year": int(year),
                    "Quarter Label": qlabel,
                    "Value": v,
                    "Activity Responsible Org": org,
                }
            )

    wb.close()
    return pd.DataFrame(rows)


def read_f31_raw(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=11, max_col=13, values_only=True):
        activity_number = clean_text(row[2])
        if not activity_number or activity_number.lower() == "total":
            continue
        year = year_from_quarter_label(row[7])
        if year not in YEARS:
            continue

        rows.append(
            {
                "Activity Number": activity_number,
                "Activity Type": clean_text(row[4]),
                "Activity Responsible Org": clean_text(row[5]),
                "Year": int(year),
                "QPR_Funds_Obligated": clean_num(row[8]) or 0.0,
                "QPR_Fund_Expended": clean_num(row[9]) or 0.0,
                "QPR_Grant_Disbursed": clean_num(row[10]) or 0.0,
                "QPR_Activity_Program_Income_Disbursed": clean_num(row[11]) or 0.0,
                "QPR_Activity_Program_Income_Received": clean_num(row[12]) or 0.0,
            }
        )

    wb.close()
    return pd.DataFrame(rows)


def build_p31_independent(p31_raw: pd.DataFrame) -> pd.DataFrame:
    agg = (
        p31_raw.groupby(
            ["Activity Number", "Activity Type", "Activity Measure Type", "Metrics", "Year"],
            dropna=False,
            as_index=False,
        )["Value"]
        .sum()
        .rename(columns={"Value": "Annual Total"})
    )
    agg["Composite Column"] = agg.apply(
        lambda r: (
            f"{sanitize_token(r['Activity Type'])}_"
            f"{sanitize_token(r['Activity Measure Type'])}_"
            f"{sanitize_token(r['Metrics'])}_"
            f"{int(r['Year'])}"
        ),
        axis=1,
    )

    wide = (
        agg.pivot_table(
            index="Activity Number",
            columns="Composite Column",
            values="Annual Total",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )
    wide.columns.name = None
    org_map = best_org_map(p31_raw, "Activity Number", "Activity Responsible Org")
    out = org_map.merge(wide, on="Activity Number", how="left")
    return out.sort_values("Activity Number").reset_index(drop=True)


def build_f31_independent(f31_raw: pd.DataFrame) -> pd.DataFrame:
    value_cols = [
        "QPR_Funds_Obligated",
        "QPR_Fund_Expended",
        "QPR_Grant_Disbursed",
        "QPR_Activity_Program_Income_Disbursed",
        "QPR_Activity_Program_Income_Received",
    ]
    agg = (
        f31_raw.groupby(["Activity Number", "Activity Type", "Year"], dropna=False, as_index=False)[value_cols]
        .sum()
    )

    long = agg.melt(
        id_vars=["Activity Number", "Activity Type", "Year"],
        value_vars=value_cols,
        var_name="Measure",
        value_name="Value",
    )
    long["Composite Column"] = long.apply(
        lambda r: f"{int(r['Year'])}_{sanitize_token(r['Activity Type'])}_{sanitize_token(r['Measure'])}",
        axis=1,
    )

    wide = (
        long.pivot_table(
            index="Activity Number",
            columns="Composite Column",
            values="Value",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )
    wide.columns.name = None
    org_map = best_org_map(f31_raw, "Activity Number", "Activity Responsible Org")
    out = org_map.merge(wide, on="Activity Number", how="left")
    return out.sort_values("Activity Number").reset_index(drop=True)


def add_check(results: List[Check], name: str, expected: float, actual: float, notes: str, tol: float = 1e-6) -> None:
    delta = float(actual) - float(expected)
    allowed = max(tol, 1e-12 * max(abs(float(expected)), 1.0))
    status = "PASS" if abs(delta) <= allowed else "FAIL"
    results.append(Check(name, status, expected, actual, delta, notes))


def sum_numeric(df: pd.DataFrame, cols: Iterable[str]) -> float:
    cols = list(cols)
    if not cols:
        return 0.0
    return float(df[cols].fillna(0).sum().sum())


def compare_values_by_cell(
    independent_df: pd.DataFrame,
    output_df: pd.DataFrame,
    key_col: str,
    value_cols: List[str],
) -> Tuple[int, int]:
    a = independent_df[[key_col] + value_cols].copy()
    b = output_df[[key_col] + value_cols].copy()
    a[key_col] = a[key_col].astype(str).str.strip()
    b[key_col] = b[key_col].astype(str).str.strip()

    a_long = a.melt(id_vars=[key_col], value_vars=value_cols, var_name="col", value_name="ind_value")
    b_long = b.melt(id_vars=[key_col], value_vars=value_cols, var_name="col", value_name="out_value")
    joined = a_long.merge(b_long, on=[key_col, "col"], how="outer")
    joined["ind_value"] = joined["ind_value"].fillna(0.0).astype(float)
    joined["out_value"] = joined["out_value"].fillna(0.0).astype(float)
    joined["delta"] = joined["out_value"] - joined["ind_value"]
    mismatch = int((joined["delta"].abs() > 1e-6).sum())
    return len(joined), mismatch


def audit(workbook_path: Path, out_csv: Path, out_xlsx: Path) -> int:
    sources = resolve_sources(ROOT)
    p31_raw = read_p31_raw(sources["HIM1_P31"])
    f31_raw = pd.concat([read_f31_raw(sources["B17_F31"]), read_f31_raw(sources["B18_F31"])], ignore_index=True)

    p31_ind = build_p31_independent(p31_raw)
    f31_ind = build_f31_independent(f31_raw)

    p31_out = pd.read_excel(workbook_path, sheet_name="P31_Restructured")
    f31_out = pd.read_excel(workbook_path, sheet_name="F31_Restructured")
    merged_out = pd.read_excel(workbook_path, sheet_name="Merged_Activity_1to1")

    results: List[Check] = []

    # row uniqueness
    add_check(results, "P31 output unique Activity Number", len(p31_out), p31_out["Activity Number"].nunique(), "One row per activity.")
    add_check(results, "F31 output unique Activity Number", len(f31_out), f31_out["Activity Number"].nunique(), "One row per activity.")
    add_check(results, "Merged output unique Activity Number", len(merged_out), merged_out["Activity Number"].nunique(), "One row per activity.")

    # schema alignment with independent rebuild
    p31_cols = [c for c in p31_out.columns if c not in ["Activity Number", "Activity Responsible Org"]]
    f31_cols = [c for c in f31_out.columns if c not in ["Activity Number", "Activity Responsible Org"]]
    p31_ind_cols = [c for c in p31_ind.columns if c not in ["Activity Number", "Activity Responsible Org"]]
    f31_ind_cols = [c for c in f31_ind.columns if c not in ["Activity Number", "Activity Responsible Org"]]

    add_check(results, "P31 output vs independent column count", len(p31_ind.columns), len(p31_out.columns), "Column schema count should match independent build.", tol=0.0)
    add_check(results, "F31 output vs independent column count", len(f31_ind.columns), len(f31_out.columns), "Column schema count should match independent build.", tol=0.0)
    add_check(results, "P31 missing columns in output", 0, len(set(p31_ind.columns) - set(p31_out.columns)), "Output should include all independent columns.", tol=0.0)
    add_check(results, "F31 missing columns in output", 0, len(set(f31_ind.columns) - set(f31_out.columns)), "Output should include all independent columns.", tol=0.0)

    # totals: raw -> independent -> output
    add_check(results, "P31 raw total vs independent total", float(p31_raw["Value"].sum()), sum_numeric(p31_ind, p31_ind_cols), "Raw P31 values preserved in independent wide.")
    add_check(results, "P31 independent total vs output total", sum_numeric(p31_ind, p31_ind_cols), sum_numeric(p31_out, p31_cols), "Independent P31 wide matches output P31 sheet.")

    measure_cols = [
        "QPR_Funds_Obligated",
        "QPR_Fund_Expended",
        "QPR_Grant_Disbursed",
        "QPR_Activity_Program_Income_Disbursed",
        "QPR_Activity_Program_Income_Received",
    ]
    for m in measure_cols:
        add_check(
            results,
            f"F31 raw {m} total vs independent total",
            float(f31_raw[m].sum()),
            sum_numeric(f31_ind, [c for c in f31_ind_cols if c.endswith("_" + sanitize_token(m))]),
            f"Raw F31 {m} preserved in independent wide.",
        )

    add_check(results, "F31 independent total vs output total", sum_numeric(f31_ind, f31_ind_cols), sum_numeric(f31_out, f31_cols), "Independent F31 wide matches output F31 sheet.")

    # year totals
    for year in YEARS:
        add_check(
            results,
            f"P31 year {year} raw vs output",
            float(p31_raw[p31_raw["Year"] == year]["Value"].sum()),
            sum_numeric(p31_out, [c for c in p31_cols if c.endswith(f"_{year}")]),
            f"Year-level P31 total for {year}.",
        )

        for m in measure_cols:
            add_check(
                results,
                f"F31 year {year} {m} raw vs output",
                float(f31_raw[f31_raw["Year"] == year][m].sum()),
                sum_numeric(
                    f31_out,
                    [c for c in f31_cols if c.startswith(f"{year}_") and c.endswith("_" + sanitize_token(m))],
                ),
                f"Year+measure F31 total for {year}, {m}.",
            )

    # cell-level comparisons for full data-path verification
    common_p31_cols = sorted(list(set(p31_ind_cols) & set(p31_cols)))
    common_f31_cols = sorted(list(set(f31_ind_cols) & set(f31_cols)))
    p31_cells, p31_mismatch = compare_values_by_cell(p31_ind, p31_out, "Activity Number", common_p31_cols)
    f31_cells, f31_mismatch = compare_values_by_cell(f31_ind, f31_out, "Activity Number", common_f31_cols)

    add_check(results, "P31 cell-level mismatches", 0, p31_mismatch, f"Compared {p31_cells:,} activity-column cells.", tol=0.0)
    add_check(results, "F31 cell-level mismatches", 0, f31_mismatch, f"Compared {f31_cells:,} activity-column cells.", tol=0.0)

    # merge path checks
    add_check(results, "Merged retains all P31 values", sum_numeric(p31_out, p31_cols), sum_numeric(merged_out, p31_cols), "No P31 values lost/duplicated in merged sheet.")
    add_check(results, "Merged retains all F31 values", sum_numeric(f31_out, f31_cols), sum_numeric(merged_out, f31_cols), "No F31 values lost/duplicated in merged sheet.")

    overlap = len(set(p31_out["Activity Number"].astype(str)) & set(f31_out["Activity Number"].astype(str)))
    merge_counts = merged_out["_merge"].value_counts().to_dict()
    add_check(results, "Merge overlap count", overlap, int(merge_counts.get("both", 0)), "Merged _merge='both' should equal key overlap.", tol=0.0)

    if "Activity Responsible Org_P31" in merged_out.columns and "Activity Responsible Org_F31" in merged_out.columns:
        both = merged_out[merged_out["_merge"] == "both"].copy()
        org_match = (
            both["Activity Responsible Org_P31"].fillna("").astype(str).str.strip()
            == both["Activity Responsible Org_F31"].fillna("").astype(str).str.strip()
        )
        add_check(results, "Merged both rows org exact-match count", len(both), int(org_match.sum()), "For overlap rows, org assignments should match source sheets.")

    # naming rules
    bad_p31 = [c for c in p31_cols if not re.match(r"^[A-Za-z0-9_]+_[A-Za-z0-9_]+_[A-Za-z0-9_]+_\d{4}$", c)]
    bad_f31 = [c for c in f31_cols if not re.match(r"^\d{4}_[A-Za-z0-9_]+_[A-Za-z0-9_]+$", c)]
    add_check(results, "P31 column naming violations", 0, len(bad_p31), "Expected ActivityType_ActivityMeasure_Metric_Year format.", tol=0.0)
    add_check(results, "F31 column naming violations", 0, len(bad_f31), "Expected Year_ActivityType_FinancialMeasure format.", tol=0.0)

    audit_df = pd.DataFrame([c.__dict__ for c in results])
    fail_df = audit_df[audit_df["status"] == "FAIL"].copy()

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    audit_df.to_csv(out_csv, index=False)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        audit_df.to_excel(writer, index=False, sheet_name="Audit_Results")
        fail_df.to_excel(writer, index=False, sheet_name="Audit_Failures")

    fail_count = int((audit_df["status"] == "FAIL").sum())
    pass_count = int((audit_df["status"] == "PASS").sum())
    print(f"Audited workbook: {workbook_path}")
    print(f"Checks -> PASS={pass_count}, FAIL={fail_count}")
    print(f"Audit CSV: {out_csv}")
    print(f"Audit XLSX: {out_xlsx}")
    return 1 if fail_count else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Independent lineage audit against original XLSX files.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Workbook to audit.")
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_AUDIT_CSV, help="Audit CSV output path.")
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_AUDIT_XLSX, help="Audit XLSX output path.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return audit(args.workbook, args.output_csv, args.output_xlsx)


if __name__ == "__main__":
    raise SystemExit(main())
