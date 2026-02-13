#!/usr/bin/env python3
"""
Build a crosswalk of activity IDs across HIM1 QPR spreadsheet exports.

Outputs:
  - outputs/model_ready/long/him1_qpr_activity_crosswalk.csv
  - outputs/model_ready/long/him1_qpr_activity_crosswalk_raw_ids.csv
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

DEFAULT_FILES = {
    "A32": "P-17-TX-48-HIM1_A32 - QPR - Activity Progress Narratives_02032026.xlsx",
    "P31": "P-17-TX-48-HIM1_P31 - QPR - Actual Accomplishments by Quarter_02032026.xlsx",
    "P33": "P-17-TX-48-HIM1_P33 - QPR - Household Characteristics for Direct Benefit Activities by Tenure and Ethnicity_02032026.xlsx",
    "B17_F31": "B-17-DM-48-0001_F31 - QPR - Fin Data by Project, Activity and Quarter_02032026.xlsx",
    "B18_F31": "B-18-DP-48-0001_F31 - QPR - Fin Data by Project, Activity and Quarter_02032026.xlsx",
}

DEFAULT_OUT_CROSSWALK = ROOT / "outputs" / "model_ready" / "long" / "him1_qpr_activity_crosswalk.csv"
DEFAULT_OUT_RAW = ROOT / "outputs" / "model_ready" / "long" / "him1_qpr_activity_crosswalk_raw_ids.csv"

TRAILING_TIMESTAMP_RE = re.compile(r"(?:[-_]\d{8,14})+$")


def _nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def normalize_activity_id(value: Any) -> Optional[str]:
    if not _nonempty(value):
        return None
    raw = str(value).replace("\xa0", " ").strip()
    raw = re.sub(r"\s+", " ", raw)
    raw = TRAILING_TIMESTAMP_RE.sub("", raw)
    normalized = raw.upper().strip()
    return normalized if normalized else None


def canonical_date(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return None


def choose_one(values: Iterable[str]) -> Optional[str]:
    cleaned = sorted(v for v in values if _nonempty(v))
    return cleaned[0] if cleaned else None


@dataclass
class SourceActivityStats:
    rows: int = 0
    raw_ids: Set[str] = field(default_factory=set)
    titles: Set[str] = field(default_factory=set)
    types: Set[str] = field(default_factory=set)
    orgs: Set[str] = field(default_factory=set)
    quarters: Set[str] = field(default_factory=set)
    dates: Set[str] = field(default_factory=set)

    def add(
        self,
        raw_id: str,
        title: Any = None,
        activity_type: Any = None,
        org: Any = None,
        quarter: Any = None,
        quarter_date: Any = None,
    ) -> None:
        self.rows += 1
        self.raw_ids.add(str(raw_id).strip())
        if _nonempty(title):
            self.titles.add(str(title).strip())
        if _nonempty(activity_type):
            self.types.add(str(activity_type).strip())
        if _nonempty(org):
            self.orgs.add(str(org).strip())
        if _nonempty(quarter):
            self.quarters.add(str(quarter).strip())
        date_iso = canonical_date(quarter_date)
        if date_iso:
            self.dates.add(date_iso)

    @property
    def date_min(self) -> Optional[str]:
        return min(self.dates) if self.dates else None

    @property
    def date_max(self) -> Optional[str]:
        return max(self.dates) if self.dates else None

    @property
    def quarter_min(self) -> Optional[str]:
        return min(self.quarters) if self.quarters else None

    @property
    def quarter_max(self) -> Optional[str]:
        return max(self.quarters) if self.quarters else None

    def raw_ids_joined(self) -> str:
        return "|".join(sorted(self.raw_ids))


def resolve_file(input_root: Path, expected_name: str) -> Path:
    direct = input_root / expected_name
    if direct.exists():
        return direct

    base = expected_name[:-5]  # strip .xlsx
    pattern = f"{base}[[]*].xlsx"
    candidates = sorted(input_root.glob(pattern))
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"Could not locate workbook for: {expected_name}")


def open_sheet(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    return wb.worksheets[0]


def build_source_maps(input_root: Path, file_map: Dict[str, str]) -> Tuple[Dict[str, Dict[str, SourceActivityStats]], Dict[str, Counter]]:
    source_stats: Dict[str, Dict[str, SourceActivityStats]] = {}
    raw_counters: Dict[str, Counter] = {}

    # A32
    ws = open_sheet(resolve_file(input_root, file_map["A32"]))
    stats: Dict[str, SourceActivityStats] = {}
    raw_counts = Counter()
    for row in ws.iter_rows(min_row=9, max_col=7, values_only=True):
        raw_id = row[0]
        canonical = normalize_activity_id(raw_id)
        if not canonical:
            continue
        raw_id_str = str(raw_id).strip()
        raw_counts[raw_id_str] += 1
        bucket = stats.setdefault(canonical, SourceActivityStats())
        bucket.add(
            raw_id=raw_id_str,
            title=row[1],
            activity_type=row[2],
            org=row[3],
            quarter_date=row[4],
            quarter=row[5],
        )
    source_stats["A32"] = stats
    raw_counters["A32"] = raw_counts

    # P31
    ws = open_sheet(resolve_file(input_root, file_map["P31"]))
    stats = {}
    raw_counts = Counter()
    for row in ws.iter_rows(min_row=10, max_col=30, values_only=True):
        raw_id = row[2]
        canonical = normalize_activity_id(raw_id)
        if not canonical:
            continue
        raw_id_str = str(raw_id).strip()
        raw_counts[raw_id_str] += 1
        bucket = stats.setdefault(canonical, SourceActivityStats())
        bucket.add(
            raw_id=raw_id_str,
            title=row[3],
            activity_type=row[0],
            org=row[1],
        )
    source_stats["P31"] = stats
    raw_counters["P31"] = raw_counts

    # P33
    ws = open_sheet(resolve_file(input_root, file_map["P33"]))
    stats = {}
    raw_counts = Counter()
    for row in ws.iter_rows(min_row=11, max_col=97, values_only=True):
        raw_id = row[1]
        canonical = normalize_activity_id(raw_id)
        if not canonical:
            continue
        raw_id_str = str(raw_id).strip()
        raw_counts[raw_id_str] += 1
        bucket = stats.setdefault(canonical, SourceActivityStats())
        bucket.add(
            raw_id=raw_id_str,
            title=row[2],
            activity_type=row[0],
            org=row[4],
        )
    source_stats["P33"] = stats
    raw_counters["P33"] = raw_counts

    # B17 F31
    ws = open_sheet(resolve_file(input_root, file_map["B17_F31"]))
    stats = {}
    raw_counts = Counter()
    for row in ws.iter_rows(min_row=11, max_col=13, values_only=True):
        raw_id = row[2]
        canonical = normalize_activity_id(raw_id)
        if not canonical:
            continue
        raw_id_str = str(raw_id).strip()
        raw_counts[raw_id_str] += 1
        bucket = stats.setdefault(canonical, SourceActivityStats())
        bucket.add(
            raw_id=raw_id_str,
            title=row[3],
            activity_type=row[4],
            org=row[5],
            quarter_date=row[6],
            quarter=row[7],
        )
    source_stats["B17_F31"] = stats
    raw_counters["B17_F31"] = raw_counts

    # B18 F31
    ws = open_sheet(resolve_file(input_root, file_map["B18_F31"]))
    stats = {}
    raw_counts = Counter()
    for row in ws.iter_rows(min_row=11, max_col=13, values_only=True):
        raw_id = row[2]
        canonical = normalize_activity_id(raw_id)
        if not canonical:
            continue
        raw_id_str = str(raw_id).strip()
        raw_counts[raw_id_str] += 1
        bucket = stats.setdefault(canonical, SourceActivityStats())
        bucket.add(
            raw_id=raw_id_str,
            title=row[3],
            activity_type=row[4],
            org=row[5],
            quarter_date=row[6],
            quarter=row[7],
        )
    source_stats["B18_F31"] = stats
    raw_counters["B18_F31"] = raw_counts

    return source_stats, raw_counters


def source_present(stats: Dict[str, Dict[str, SourceActivityStats]], source: str, canonical_id: str) -> bool:
    return canonical_id in stats.get(source, {})


def source_rows(stats: Dict[str, Dict[str, SourceActivityStats]], source: str, canonical_id: str) -> int:
    activity = stats.get(source, {}).get(canonical_id)
    return activity.rows if activity else 0


def source_activity(stats: Dict[str, Dict[str, SourceActivityStats]], source: str, canonical_id: str) -> Optional[SourceActivityStats]:
    return stats.get(source, {}).get(canonical_id)


def build_crosswalk_rows(source_stats: Dict[str, Dict[str, SourceActivityStats]]) -> List[Dict[str, Any]]:
    all_ids: Set[str] = set()
    for per_source in source_stats.values():
        all_ids.update(per_source.keys())

    rows: List[Dict[str, Any]] = []
    for canonical_id in sorted(all_ids):
        a32 = source_activity(source_stats, "A32", canonical_id)
        p31 = source_activity(source_stats, "P31", canonical_id)
        p33 = source_activity(source_stats, "P33", canonical_id)
        b17 = source_activity(source_stats, "B17_F31", canonical_id)
        b18 = source_activity(source_stats, "B18_F31", canonical_id)

        in_a32 = a32 is not None
        in_p31 = p31 is not None
        in_p33 = p33 is not None
        in_b17 = b17 is not None
        in_b18 = b18 is not None
        in_b_any = in_b17 or in_b18
        source_count = int(in_a32) + int(in_p31) + int(in_p33) + int(in_b17) + int(in_b18)

        rows.append(
            {
                "canonical_activity_id": canonical_id,
                "source_count": source_count,
                "in_a32": int(in_a32),
                "in_p31": int(in_p31),
                "in_p33": int(in_p33),
                "in_b17_f31": int(in_b17),
                "in_b18_f31": int(in_b18),
                "in_b_any_f31": int(in_b_any),
                "a32_row_count": source_rows(source_stats, "A32", canonical_id),
                "p31_row_count": source_rows(source_stats, "P31", canonical_id),
                "p33_row_count": source_rows(source_stats, "P33", canonical_id),
                "b17_f31_row_count": source_rows(source_stats, "B17_F31", canonical_id),
                "b18_f31_row_count": source_rows(source_stats, "B18_F31", canonical_id),
                "b_any_f31_row_count": source_rows(source_stats, "B17_F31", canonical_id)
                + source_rows(source_stats, "B18_F31", canonical_id),
                "a32_quarter_count": len(a32.quarters) if a32 else 0,
                "a32_date_min": a32.date_min if a32 else None,
                "a32_date_max": a32.date_max if a32 else None,
                "b17_f31_quarter_count": len(b17.quarters) if b17 else 0,
                "b17_f31_date_min": b17.date_min if b17 else None,
                "b17_f31_date_max": b17.date_max if b17 else None,
                "b18_f31_quarter_count": len(b18.quarters) if b18 else 0,
                "b18_f31_date_min": b18.date_min if b18 else None,
                "b18_f31_date_max": b18.date_max if b18 else None,
                "a32_title_example": choose_one(a32.titles) if a32 else None,
                "p31_title_example": choose_one(p31.titles) if p31 else None,
                "p33_title_example": choose_one(p33.titles) if p33 else None,
                "b17_title_example": choose_one(b17.titles) if b17 else None,
                "b18_title_example": choose_one(b18.titles) if b18 else None,
                "a32_type_example": choose_one(a32.types) if a32 else None,
                "p31_type_example": choose_one(p31.types) if p31 else None,
                "p33_type_example": choose_one(p33.types) if p33 else None,
                "b17_type_example": choose_one(b17.types) if b17 else None,
                "b18_type_example": choose_one(b18.types) if b18 else None,
                "a32_org_example": choose_one(a32.orgs) if a32 else None,
                "p31_org_example": choose_one(p31.orgs) if p31 else None,
                "p33_org_example": choose_one(p33.orgs) if p33 else None,
                "b17_org_example": choose_one(b17.orgs) if b17 else None,
                "b18_org_example": choose_one(b18.orgs) if b18 else None,
                "raw_ids_a32": a32.raw_ids_joined() if a32 else None,
                "raw_ids_p31": p31.raw_ids_joined() if p31 else None,
                "raw_ids_p33": p33.raw_ids_joined() if p33 else None,
                "raw_ids_b17_f31": b17.raw_ids_joined() if b17 else None,
                "raw_ids_b18_f31": b18.raw_ids_joined() if b18 else None,
            }
        )
    return rows


def build_raw_rows(raw_counters: Dict[str, Counter]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for source in sorted(raw_counters):
        for raw_id, row_count in sorted(raw_counters[source].items()):
            rows.append(
                {
                    "source": source,
                    "raw_activity_id": raw_id,
                    "canonical_activity_id": normalize_activity_id(raw_id),
                    "row_count": row_count,
                }
            )
    return rows


def write_csv(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"No rows to write for {path}")
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build HIM1 QPR activity crosswalk CSVs.")
    parser.add_argument(
        "--input-root",
        type=Path,
        default=ROOT,
        help="Directory containing the QPR XLSX exports.",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=DEFAULT_OUT_CROSSWALK,
        help="Output CSV path for canonical activity crosswalk.",
    )
    parser.add_argument(
        "--output-raw-csv",
        type=Path,
        default=DEFAULT_OUT_RAW,
        help="Output CSV path for raw-to-canonical activity IDs.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_stats, raw_counters = build_source_maps(args.input_root, DEFAULT_FILES)
    crosswalk_rows = build_crosswalk_rows(source_stats)
    raw_rows = build_raw_rows(raw_counters)
    write_csv(args.output_csv, crosswalk_rows)
    write_csv(args.output_raw_csv, raw_rows)
    print(f"Wrote {len(crosswalk_rows)} crosswalk rows -> {args.output_csv}")
    print(f"Wrote {len(raw_rows)} raw-id rows -> {args.output_raw_csv}")


if __name__ == "__main__":
    main()
