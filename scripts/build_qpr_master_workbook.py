#!/usr/bin/env python3
"""
Build a linked master workbook from the 8 canonical QPR XLSX exports.

This script normalizes report-style sheets (F31, F33, A31, A32, P31, P33),
builds link/audit tables using keys present in the reports, and writes a
single master Excel workbook.
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

SOURCE_SPECS = [
    {
        "key": "B17_F31",
        "prefix": "B-17-DM-48-0001_F31",
        "grant": "B-17-DM-48-0001",
        "report": "F31",
    },
    {
        "key": "B17_F33",
        "prefix": "B-17-DM-48-0001_F33",
        "grant": "B-17-DM-48-0001",
        "report": "F33",
    },
    {
        "key": "B18_F31",
        "prefix": "B-18-DP-48-0001_F31",
        "grant": "B-18-DP-48-0001",
        "report": "F31",
    },
    {
        "key": "B18_F33",
        "prefix": "B-18-DP-48-0001_F33",
        "grant": "B-18-DP-48-0001",
        "report": "F33",
    },
    {
        "key": "HIM1_A31",
        "prefix": "P-17-TX-48-HIM1_A31",
        "grant": "P-17-TX-48-HIM1",
        "report": "A31",
    },
    {
        "key": "HIM1_A32",
        "prefix": "P-17-TX-48-HIM1_A32",
        "grant": "P-17-TX-48-HIM1",
        "report": "A32",
    },
    {
        "key": "HIM1_P31",
        "prefix": "P-17-TX-48-HIM1_P31",
        "grant": "P-17-TX-48-HIM1",
        "report": "P31",
    },
    {
        "key": "HIM1_P33",
        "prefix": "P-17-TX-48-HIM1_P33",
        "grant": "P-17-TX-48-HIM1",
        "report": "P33",
    },
]

TRAILING_TIMESTAMP_RE = re.compile(r"(?:[-_]\d{8,14})+$")
QUARTER_RE = re.compile(r"(\d{4})\s*Q([1-4])", re.IGNORECASE)
BRACKETED_COPY_RE = re.compile(r"\[\d+\](?=\.xlsx$)", re.IGNORECASE)


def clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).replace("\xa0", " ").strip()
    return s if s else None


def clean_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Preserve zeros; guard NaN.
        return float(value) if value == value else None
    s = str(value).strip()
    if not s:
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    s = s.replace(",", "").replace("$", "").strip()
    if s in {"", "-", "--"}:
        return None
    try:
        num = float(s)
        return -num if negative else num
    except ValueError:
        return None


def date_iso(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = clean_str(value)
    if not s:
        return None
    s = s.replace("T", " ")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    if len(s) >= 10:
        stub = s[:10]
        try:
            return datetime.strptime(stub, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return None
    return None


def quarter_label(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        quarter = (value.month - 1) // 3 + 1
        return f"{value.year} Q{quarter}"
    s = clean_str(value)
    if not s:
        return None
    m = QUARTER_RE.search(s)
    if m:
        return f"{m.group(1)} Q{m.group(2)}"
    iso = date_iso(s)
    if iso:
        dt = datetime.strptime(iso, "%Y-%m-%d")
        quarter = (dt.month - 1) // 3 + 1
        return f"{dt.year} Q{quarter}"
    return None


def canonical_activity_id(value: Any) -> Optional[str]:
    raw = clean_str(value)
    if not raw:
        return None
    raw = re.sub(r"\s+", " ", raw)
    raw = TRAILING_TIMESTAMP_RE.sub("", raw)
    out = raw.upper().strip()
    return out if out else None


def canonical_project_title(value: Any) -> Optional[str]:
    raw = clean_str(value)
    if not raw:
        return None
    norm = re.sub(r"[^A-Z0-9]+", " ", raw.upper()).strip()
    norm = re.sub(r"\s+", " ", norm)
    return norm if norm else None


def normalize_for_match(value: Any) -> str:
    raw = clean_str(value)
    if not raw:
        return ""
    norm = re.sub(r"[^A-Z0-9]+", " ", raw.upper()).strip()
    return re.sub(r"\s+", " ", norm)


def project_aliases(project_title: str) -> List[str]:
    base = normalize_for_match(project_title)
    if not base:
        return []
    aliases = {base}

    if base.endswith(" PROGRAM"):
        aliases.add(base[: -len(" PROGRAM")].strip())

    for prefix in ("STATE MANAGED ", "STATE MI ", "LOCAL "):
        if base.startswith(prefix):
            stripped = base[len(prefix) :].strip()
            if stripped:
                aliases.add(stripped)
                if stripped.endswith(" PROGRAM"):
                    aliases.add(stripped[: -len(" PROGRAM")].strip())

    if "/" in base:
        aliases.add(base.replace("/", " "))

    if "CITY OF HOUSTON" in base:
        aliases.add("CITY OF HOUSTON")
        aliases.add("HOUSTON")
    if "HARRIS COUNTY" in base:
        aliases.add("HARRIS COUNTY")
    if "BUYOUT" in base:
        aliases.add("BUYOUT")
        aliases.add("BUYOUT PROGRAM")
    if "INFRASTRUCTURE" in base:
        aliases.add("INFRASTRUCTURE")
        aliases.add("INFRASTRUCTURE PROGRAM")
    if "REIMBURSEMENT" in base:
        aliases.add("REIMBURSEMENT")
        aliases.add("REIMBURSEMENT PROGRAM")
    if "AFFORDABLE RENTAL" in base:
        aliases.add("AFFORDABLE RENTAL")
    if "ECONOMIC REVITALIZATION" in base:
        aliases.add("ECONOMIC REVITALIZATION")
    if "PREPS" in base:
        aliases.add("PREPS")

    generic = {"PROGRAM", "STATE", "LOCAL", "MANAGED", "CITY", "COUNTY"}
    cleaned = [a for a in aliases if a and a not in generic]
    return sorted(cleaned, key=lambda x: (-len(x), x))


def preferred_candidate(paths: Sequence[Path]) -> Path:
    def rank(p: Path) -> Tuple[int, int, str]:
        has_bracket_copy = 1 if BRACKETED_COPY_RE.search(p.name) else 0
        return (has_bracket_copy, len(p.name), p.name)

    return sorted(paths, key=rank)[0]


def resolve_source_files(input_root: Path) -> Dict[str, Path]:
    candidates = [
        p
        for p in input_root.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("._")
    ]
    resolved: Dict[str, Path] = {}
    for spec in SOURCE_SPECS:
        prefix = spec["prefix"]
        matches = [p for p in candidates if p.name.startswith(prefix)]
        if not matches:
            raise FileNotFoundError(f"Missing source workbook for prefix: {prefix}")
        resolved[spec["key"]] = preferred_candidate(matches)
    return resolved


def open_first_sheet(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    return wb, wb.worksheets[0]


def parse_f31(path: Path, source_key: str, grant_number: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=11, max_col=13, values_only=True):
        activity_number = clean_str(row[2])
        if not activity_number or activity_number.lower() == "total":
            continue
        quarter = quarter_label(row[7]) or quarter_label(row[6])
        if not quarter:
            continue
        project_title = clean_str(row[1])
        rows.append(
            {
                "source_key": source_key,
                "source_file": path.name,
                "grant_number": grant_number,
                "project_number": clean_str(row[0]),
                "project_title": project_title,
                "project_title_norm": canonical_project_title(project_title),
                "activity_number": activity_number,
                "canonical_activity_id": canonical_activity_id(activity_number),
                "activity_title": clean_str(row[3]),
                "activity_type": clean_str(row[4]),
                "responsible_org": clean_str(row[5]),
                "quarter_start_date": date_iso(row[6]),
                "quarter_label": quarter,
                "obligated_usd": clean_float(row[8]),
                "expended_usd": clean_float(row[9]),
                "disbursed_usd": clean_float(row[10]),
                "program_income_disbursed_usd": clean_float(row[11]),
                "program_income_received_usd": clean_float(row[12]),
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def parse_f33(path: Path, source_key: str, grant_number: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    header_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]
    quarter_row = list(ws.iter_rows(min_row=10, max_row=10, values_only=True))[0]

    quarter_cols: List[Tuple[int, str, Optional[str]]] = []
    for idx in range(3, len(quarter_row)):
        qlabel = quarter_label(quarter_row[idx]) or quarter_label(header_row[idx])
        if not qlabel:
            continue
        quarter_cols.append((idx, qlabel, date_iso(header_row[idx])))

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        project_title = clean_str(row[0]) if len(row) > 0 else None
        if not project_title:
            continue
        project_number = clean_str(row[1]) if len(row) > 1 else None
        for idx, qlabel, qstart in quarter_cols:
            if idx >= len(row):
                continue
            value = clean_float(row[idx])
            if value is None:
                continue
            rows.append(
                {
                    "source_key": source_key,
                    "source_file": path.name,
                    "grant_number": grant_number,
                    "project_number": project_number,
                    "project_title": project_title,
                    "project_title_norm": canonical_project_title(project_title),
                    "quarter_start_date": qstart,
                    "quarter_label": qlabel,
                    "f33_value_usd": value,
                }
            )
    wb.close()
    return pd.DataFrame(rows)


def parse_a31(path: Path, source_key: str, default_grant: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=6, max_col=6, values_only=True):
        grant_number = clean_str(row[0]) or default_grant
        quarter = quarter_label(row[4]) or quarter_label(row[3])
        if not quarter:
            continue
        rows.append(
            {
                "source_key": source_key,
                "source_file": path.name,
                "grant_number": grant_number,
                "grantee_state": clean_str(row[1]),
                "grantee_name": clean_str(row[2]),
                "quarter_start_date": date_iso(row[3]),
                "quarter_label": quarter,
                "executive_summary": clean_str(row[5]),
                "summary_char_count": len(clean_str(row[5]) or ""),
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def parse_a32(path: Path, source_key: str, grant_number: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=9, max_col=7, values_only=True):
        activity_number = clean_str(row[0])
        if not activity_number:
            continue
        quarter = quarter_label(row[5]) or quarter_label(row[4])
        if not quarter:
            continue
        narrative = clean_str(row[6])
        rows.append(
            {
                "source_key": source_key,
                "source_file": path.name,
                "grant_number": grant_number,
                "activity_number": activity_number,
                "canonical_activity_id": canonical_activity_id(activity_number),
                "activity_title": clean_str(row[1]),
                "activity_type": clean_str(row[2]),
                "responsible_org": clean_str(row[3]),
                "quarter_start_date": date_iso(row[4]),
                "quarter_label": quarter,
                "activity_narrative": narrative,
                "narrative_char_count": len(narrative or ""),
                "has_payroll_keyword": int("payroll" in (narrative or "").lower()),
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def parse_p31(path: Path, source_key: str, grant_number: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    header_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    quarter_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    quarter_cols: List[Tuple[int, str, Optional[str]]] = []
    for idx in range(6, len(quarter_row)):
        qlabel = quarter_label(quarter_row[idx]) or quarter_label(header_row[idx])
        if not qlabel:
            continue
        quarter_cols.append((idx, qlabel, date_iso(header_row[idx])))

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        activity_number = clean_str(row[2]) if len(row) > 2 else None
        if not activity_number:
            continue
        for idx, qlabel, qstart in quarter_cols:
            if idx >= len(row):
                continue
            actual = clean_float(row[idx])
            if actual is None:
                continue
            rows.append(
                {
                    "source_key": source_key,
                    "source_file": path.name,
                    "grant_number": grant_number,
                    "activity_number": activity_number,
                    "canonical_activity_id": canonical_activity_id(activity_number),
                    "activity_title": clean_str(row[3]) if len(row) > 3 else None,
                    "activity_type": clean_str(row[0]) if len(row) > 0 else None,
                    "responsible_org": clean_str(row[1]) if len(row) > 1 else None,
                    "measure_type": clean_str(row[4]) if len(row) > 4 else None,
                    "metric_label": clean_str(row[5]) if len(row) > 5 else None,
                    "quarter_start_date": qstart,
                    "quarter_label": qlabel,
                    "actual_value": actual,
                }
            )
    wb.close()
    return pd.DataFrame(rows)


def parse_p33(path: Path, source_key: str, grant_number: str) -> pd.DataFrame:
    wb, ws = open_first_sheet(path)
    cat_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    demo_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    col_map: List[Tuple[int, str, str]] = []
    current_category = ""
    for idx in range(5, min(len(cat_row), len(demo_row))):
        category = clean_str(cat_row[idx])
        if category:
            current_category = category
        demographic_group = clean_str(demo_row[idx])
        if current_category and demographic_group:
            col_map.append((idx, current_category, demographic_group))

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        activity_number = clean_str(row[1]) if len(row) > 1 else None
        if not activity_number:
            continue
        for idx, category, demographic_group in col_map:
            if idx >= len(row):
                continue
            value = clean_float(row[idx])
            if value is None:
                continue
            rows.append(
                {
                    "source_key": source_key,
                    "source_file": path.name,
                    "grant_number": grant_number,
                    "activity_number": activity_number,
                    "canonical_activity_id": canonical_activity_id(activity_number),
                    "activity_title": clean_str(row[2]) if len(row) > 2 else None,
                    "activity_type": clean_str(row[0]) if len(row) > 0 else None,
                    "national_objective": clean_str(row[3]) if len(row) > 3 else None,
                    "responsible_org": clean_str(row[4]) if len(row) > 4 else None,
                    "beneficiary_category": category,
                    "demographic_group": demographic_group,
                    "value": value,
                }
            )
    wb.close()
    return pd.DataFrame(rows)


def first_nonempty(values: Iterable[Any]) -> Optional[str]:
    for value in values:
        s = clean_str(value)
        if s:
            return s
    return None


def build_activity_link(
    f31_df: pd.DataFrame, a32_df: pd.DataFrame, p31_df: pd.DataFrame, p33_df: pd.DataFrame
) -> pd.DataFrame:
    all_ids = set()
    for df in (f31_df, a32_df, p31_df, p33_df):
        if df.empty:
            continue
        all_ids.update(df["canonical_activity_id"].dropna().unique().tolist())

    rows: List[Dict[str, Any]] = []
    for activity_id in sorted(all_ids):
        f31_rows = f31_df[f31_df["canonical_activity_id"] == activity_id]
        a32_rows = a32_df[a32_df["canonical_activity_id"] == activity_id]
        p31_rows = p31_df[p31_df["canonical_activity_id"] == activity_id]
        p33_rows = p33_df[p33_df["canonical_activity_id"] == activity_id]

        in_f31 = int(not f31_rows.empty)
        in_a32 = int(not a32_rows.empty)
        in_p31 = int(not p31_rows.empty)
        in_p33 = int(not p33_rows.empty)
        source_count = in_f31 + in_a32 + in_p31 + in_p33

        if source_count >= 3:
            tier = "high"
        elif source_count == 2:
            tier = "medium"
        else:
            tier = "low"

        rows.append(
            {
                "canonical_activity_id": activity_id,
                "source_count": source_count,
                "link_strength": tier,
                "in_f31": in_f31,
                "in_a32": in_a32,
                "in_p31": in_p31,
                "in_p33": in_p33,
                "f31_rows": len(f31_rows),
                "a32_rows": len(a32_rows),
                "p31_rows": len(p31_rows),
                "p33_rows": len(p33_rows),
                "f31_quarter_count": f31_rows["quarter_label"].nunique() if not f31_rows.empty else 0,
                "a32_quarter_count": a32_rows["quarter_label"].nunique() if not a32_rows.empty else 0,
                "p31_quarter_count": p31_rows["quarter_label"].nunique() if not p31_rows.empty else 0,
                "title_example_f31": first_nonempty(f31_rows["activity_title"]) if not f31_rows.empty else None,
                "title_example_a32": first_nonempty(a32_rows["activity_title"]) if not a32_rows.empty else None,
                "title_example_p31": first_nonempty(p31_rows["activity_title"]) if not p31_rows.empty else None,
                "title_example_p33": first_nonempty(p33_rows["activity_title"]) if not p33_rows.empty else None,
            }
        )

    out = pd.DataFrame(rows)
    return out.sort_values(["source_count", "canonical_activity_id"], ascending=[False, True]).reset_index(drop=True)


def build_activity_quarter_link(
    f31_df: pd.DataFrame, a32_df: pd.DataFrame, p31_df: pd.DataFrame
) -> pd.DataFrame:
    f31_counts = (
        f31_df.groupby(["canonical_activity_id", "quarter_label", "source_key"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    for col in ("B17_F31", "B18_F31"):
        if col not in f31_counts.columns:
            f31_counts[col] = 0
    f31_counts["f31_rows"] = f31_counts["B17_F31"] + f31_counts["B18_F31"]

    a32_counts = (
        a32_df.groupby(["canonical_activity_id", "quarter_label"])
        .size()
        .rename("a32_rows")
        .reset_index()
    )
    p31_counts = (
        p31_df.groupby(["canonical_activity_id", "quarter_label"])
        .size()
        .rename("p31_rows")
        .reset_index()
    )

    out = f31_counts.merge(a32_counts, how="outer", on=["canonical_activity_id", "quarter_label"])
    out = out.merge(p31_counts, how="outer", on=["canonical_activity_id", "quarter_label"])
    out = out.fillna(0)
    for col in ("B17_F31", "B18_F31", "f31_rows", "a32_rows", "p31_rows"):
        out[col] = out[col].astype(int)
    out["source_count"] = (
        (out["f31_rows"] > 0).astype(int)
        + (out["a32_rows"] > 0).astype(int)
        + (out["p31_rows"] > 0).astype(int)
    )
    out["link_strength"] = out["source_count"].map({3: "high", 2: "medium", 1: "low"}).fillna("low")
    return out.sort_values(["source_count", "canonical_activity_id", "quarter_label"], ascending=[False, True, True]).reset_index(drop=True)


def build_narrative_project_mentions(
    a31_df: pd.DataFrame,
    a32_df: pd.DataFrame,
    f33_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    project_rows = (
        f33_df[["project_title", "project_title_norm"]]
        .dropna(subset=["project_title_norm"])
        .drop_duplicates()
        .sort_values("project_title_norm")
    )

    narrative_rows: List[Dict[str, Any]] = []
    for idx, row in a31_df.reset_index(drop=True).iterrows():
        text = " ".join(
            part
            for part in [
                clean_str(row.get("grantee_name")),
                clean_str(row.get("executive_summary")),
            ]
            if part
        )
        narrative_rows.append(
            {
                "quarter_label": clean_str(row.get("quarter_label")),
                "narrative_source": "A31",
                "source_row_id": f"A31:{idx + 1}",
                "text_norm": normalize_for_match(text),
            }
        )

    for idx, row in a32_df.reset_index(drop=True).iterrows():
        text = " ".join(
            part
            for part in [
                clean_str(row.get("activity_title")),
                clean_str(row.get("responsible_org")),
                clean_str(row.get("activity_narrative")),
            ]
            if part
        )
        narrative_rows.append(
            {
                "quarter_label": clean_str(row.get("quarter_label")),
                "narrative_source": "A32",
                "source_row_id": f"A32:{idx + 1}",
                "text_norm": normalize_for_match(text),
            }
        )

    narrative_df = pd.DataFrame(narrative_rows)
    if narrative_df.empty:
        empty_mentions = pd.DataFrame(
            columns=[
                "quarter_label",
                "project_title_norm",
                "project_title",
                "narrative_mention_count",
                "narrative_source_count",
                "narrative_sources",
                "alias_examples",
            ]
        )
        empty_quarters = pd.DataFrame(columns=["quarter_label", "narrative_row_count"])
        return empty_mentions, empty_quarters

    narrative_quarters_df = (
        narrative_df.groupby("quarter_label")
        .size()
        .rename("narrative_row_count")
        .reset_index()
        .sort_values("quarter_label")
        .reset_index(drop=True)
    )

    mention_rows: List[Dict[str, Any]] = []
    records = narrative_df.to_dict("records")
    for _, prow in project_rows.iterrows():
        project_title = clean_str(prow["project_title"])
        project_title_norm = clean_str(prow["project_title_norm"])
        if not project_title_norm:
            continue
        aliases = project_aliases(project_title or project_title_norm)
        if not aliases:
            continue
        for nrow in records:
            quarter = clean_str(nrow["quarter_label"])
            text_norm = clean_str(nrow["text_norm"]) or ""
            if not quarter or not text_norm:
                continue
            matched_aliases = [alias for alias in aliases if alias in text_norm]
            if not matched_aliases:
                continue
            mention_rows.append(
                {
                    "quarter_label": quarter,
                    "project_title_norm": project_title_norm,
                    "project_title": project_title,
                    "narrative_source": nrow["narrative_source"],
                    "source_row_id": nrow["source_row_id"],
                    "matched_alias": matched_aliases[0],
                }
            )

    if not mention_rows:
        empty_mentions = pd.DataFrame(
            columns=[
                "quarter_label",
                "project_title_norm",
                "project_title",
                "narrative_mention_count",
                "narrative_source_count",
                "narrative_sources",
                "alias_examples",
            ]
        )
        return empty_mentions, narrative_quarters_df

    mentions_df = pd.DataFrame(mention_rows)
    mentions_agg_df = (
        mentions_df.groupby(["quarter_label", "project_title_norm", "project_title"], dropna=False)
        .agg(
            narrative_mention_count=("source_row_id", "nunique"),
            narrative_source_count=("narrative_source", "nunique"),
            narrative_sources=("narrative_source", lambda s: "|".join(sorted(set(s)))),
            alias_examples=("matched_alias", lambda s: "|".join(sorted(set(s))[:5])),
        )
        .reset_index()
        .sort_values(["quarter_label", "project_title_norm"])
        .reset_index(drop=True)
    )
    return mentions_agg_df, narrative_quarters_df


def build_project_quarter_link(
    f31_df: pd.DataFrame,
    f33_df: pd.DataFrame,
    narrative_mentions_df: pd.DataFrame,
    narrative_quarters_df: pd.DataFrame,
) -> pd.DataFrame:
    f31_agg = (
        f31_df.groupby(["grant_number", "project_title_norm", "quarter_label"], dropna=False)
        .agg(
            f31_rows=("project_title", "size"),
            f31_obligated_usd=("obligated_usd", "sum"),
            f31_project_title=("project_title", "first"),
            f31_project_numbers=("project_number", lambda s: "|".join(sorted({v for v in s.dropna().astype(str)}))),
        )
        .reset_index()
    )

    f33_agg = (
        f33_df.groupby(["grant_number", "project_title_norm", "quarter_label"], dropna=False)
        .agg(
            f33_rows=("project_title", "size"),
            f33_value_usd=("f33_value_usd", "sum"),
            f33_project_title=("project_title", "first"),
            f33_project_numbers=("project_number", lambda s: "|".join(sorted({v for v in s.dropna().astype(str)}))),
        )
        .reset_index()
    )

    out = f31_agg.merge(
        f33_agg,
        how="outer",
        on=["grant_number", "project_title_norm", "quarter_label"],
    )
    out = out.merge(
        narrative_mentions_df[
            [
                "quarter_label",
                "project_title_norm",
                "narrative_mention_count",
                "narrative_source_count",
                "narrative_sources",
                "alias_examples",
            ]
        ],
        how="left",
        on=["quarter_label", "project_title_norm"],
    )
    out = out.merge(
        narrative_quarters_df[["quarter_label", "narrative_row_count"]],
        how="left",
        on="quarter_label",
    )
    out["f31_rows"] = out["f31_rows"].fillna(0).astype(int)
    out["f33_rows"] = out["f33_rows"].fillna(0).astype(int)
    out["narrative_mention_count"] = out["narrative_mention_count"].fillna(0).astype(int)
    out["narrative_source_count"] = out["narrative_source_count"].fillna(0).astype(int)
    out["narrative_row_count"] = out["narrative_row_count"].fillna(0).astype(int)
    out["narrative_sources"] = out["narrative_sources"].fillna("")
    out["alias_examples"] = out["alias_examples"].fillna("")
    out["in_f31"] = (out["f31_rows"] > 0).astype(int)
    out["in_f33"] = (out["f33_rows"] > 0).astype(int)
    out["source_count"] = out["in_f31"] + out["in_f33"]
    out["link_strength"] = out["source_count"].map({2: "high", 1: "low"}).fillna("low")
    out["f31_obligated_usd"] = out["f31_obligated_usd"].fillna(0.0)
    out["f33_value_usd"] = out["f33_value_usd"].fillna(0.0)
    out["amount_delta_usd"] = out["f31_obligated_usd"] - out["f33_value_usd"]
    out["narrative_project_mentioned"] = (out["narrative_mention_count"] > 0).astype(int)
    out["quarter_has_narrative"] = (out["narrative_row_count"] > 0).astype(int)
    out["is_structural_zero_gap"] = (
        (out["in_f31"] == 0) & (out["in_f33"] == 1) & (out["f33_value_usd"] == 0)
    ).astype(int)
    out["is_unresolved_nonzero_gap"] = (
        (out["in_f31"] == 0) & (out["in_f33"] == 1) & (out["f33_value_usd"] != 0)
    ).astype(int)

    out["resolution_status"] = "direct_match"
    out.loc[(out["in_f31"] == 1) & (out["in_f33"] == 0), "resolution_status"] = "f31_only_gap"
    out.loc[(out["in_f31"] == 0) & (out["in_f33"] == 1), "resolution_status"] = "f33_only_gap"
    out.loc[
        (out["in_f31"] == 0)
        & (out["in_f33"] == 1)
        & (out["is_structural_zero_gap"] == 1)
        & (out["narrative_project_mentioned"] == 1),
        "resolution_status",
    ] = "inferred_zero_from_narrative"
    out.loc[
        (out["in_f31"] == 0)
        & (out["in_f33"] == 1)
        & (out["is_structural_zero_gap"] == 1)
        & (out["narrative_project_mentioned"] == 0)
        & (out["quarter_has_narrative"] == 1),
        "resolution_status",
    ] = "inferred_zero_from_quarter_narrative"
    out.loc[
        (out["in_f31"] == 0)
        & (out["in_f33"] == 1)
        & (out["is_structural_zero_gap"] == 1)
        & (out["quarter_has_narrative"] == 0),
        "resolution_status",
    ] = "inferred_zero_no_narrative"
    out.loc[
        (out["in_f31"] == 0)
        & (out["in_f33"] == 1)
        & (out["is_unresolved_nonzero_gap"] == 1),
        "resolution_status",
    ] = "unresolved_nonzero_gap"

    out["reliable_link_flag"] = out["resolution_status"].isin(
        {
            "direct_match",
            "inferred_zero_from_narrative",
            "inferred_zero_from_quarter_narrative",
            "inferred_zero_no_narrative",
        }
    ).astype(int)
    out["narrative_backed_flag"] = out["resolution_status"].isin(
        {
            "direct_match",
            "inferred_zero_from_narrative",
            "inferred_zero_from_quarter_narrative",
        }
    ).astype(int)
    return out.sort_values(
        ["source_count", "grant_number", "project_title_norm", "quarter_label"],
        ascending=[False, True, True, True],
    ).reset_index(drop=True)


def build_quarter_timeline_link(
    f31_df: pd.DataFrame,
    f33_df: pd.DataFrame,
    a31_df: pd.DataFrame,
    a32_df: pd.DataFrame,
    p31_df: pd.DataFrame,
) -> pd.DataFrame:
    def quarter_counts(df: pd.DataFrame, value_col: str) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame(columns=["quarter_label", value_col])
        return df.groupby("quarter_label").size().rename(value_col).reset_index()

    out = quarter_counts(f31_df, "f31_rows")
    out = out.merge(quarter_counts(f33_df, "f33_rows"), how="outer", on="quarter_label")
    out = out.merge(quarter_counts(a31_df, "a31_rows"), how="outer", on="quarter_label")
    out = out.merge(quarter_counts(a32_df, "a32_rows"), how="outer", on="quarter_label")
    out = out.merge(quarter_counts(p31_df, "p31_rows"), how="outer", on="quarter_label")
    out = out.fillna(0)
    for col in ("f31_rows", "f33_rows", "a31_rows", "a32_rows", "p31_rows"):
        out[col] = out[col].astype(int)
    out["source_count"] = (
        (out["f31_rows"] > 0).astype(int)
        + (out["f33_rows"] > 0).astype(int)
        + (out["a31_rows"] > 0).astype(int)
        + (out["a32_rows"] > 0).astype(int)
        + (out["p31_rows"] > 0).astype(int)
    )
    out["link_strength"] = pd.cut(
        out["source_count"],
        bins=[-1, 1, 2, 5],
        labels=["low", "medium", "high"],
    ).astype(str)
    return out.sort_values("quarter_label").reset_index(drop=True)


def build_coverage_summary(
    activity_link_df: pd.DataFrame,
    activity_quarter_df: pd.DataFrame,
    project_quarter_df: pd.DataFrame,
    timeline_df: pd.DataFrame,
) -> pd.DataFrame:
    def pct(num: int, den: int) -> float:
        if den == 0:
            return 0.0
        return round(100.0 * num / den, 2)

    rows = []

    total = len(activity_link_df)
    matched = int((activity_link_df["source_count"] >= 2).sum()) if total else 0
    rows.append(
        {
            "link_domain": "activity_id_only",
            "total_keys": total,
            "matched_keys_ge_2_sources": matched,
            "match_rate_pct": pct(matched, total),
            "notes": "F31 + A32 + P31 + P33 on canonical activity id",
        }
    )

    total = len(activity_quarter_df)
    matched = int((activity_quarter_df["source_count"] >= 2).sum()) if total else 0
    rows.append(
        {
            "link_domain": "activity_id_plus_quarter",
            "total_keys": total,
            "matched_keys_ge_2_sources": matched,
            "match_rate_pct": pct(matched, total),
            "notes": "F31 + A32 + P31 on canonical activity id + quarter",
        }
    )

    total = len(project_quarter_df)
    matched = int((project_quarter_df["source_count"] >= 2).sum()) if total else 0
    rows.append(
        {
            "link_domain": "project_title_plus_quarter_plus_grant_direct",
            "total_keys": total,
            "matched_keys_ge_2_sources": matched,
            "match_rate_pct": pct(matched, total),
            "notes": "Direct F31 <-> F33 by grant + normalized project title + quarter",
        }
    )

    inferred = int((project_quarter_df["reliable_link_flag"] == 1).sum()) if total else 0
    rows.append(
        {
            "link_domain": "project_title_plus_quarter_plus_grant_inferred",
            "total_keys": total,
            "matched_keys_ge_2_sources": inferred,
            "match_rate_pct": pct(inferred, total),
            "notes": "Direct matches plus inferred structural-zero F33 rows",
        }
    )

    narrative_backed = int((project_quarter_df["narrative_backed_flag"] == 1).sum()) if total else 0
    rows.append(
        {
            "link_domain": "project_title_plus_quarter_plus_grant_narrative_backed",
            "total_keys": total,
            "matched_keys_ge_2_sources": narrative_backed,
            "match_rate_pct": pct(narrative_backed, total),
            "notes": "Direct matches plus structural-zero rows supported by narrative quarter/project context",
        }
    )

    total = len(timeline_df)
    matched = int((timeline_df["source_count"] >= 2).sum()) if total else 0
    rows.append(
        {
            "link_domain": "quarter_timeline",
            "total_keys": total,
            "matched_keys_ge_2_sources": matched,
            "match_rate_pct": pct(matched, total),
            "notes": "Quarter overlap across F31/F33/A31/A32/P31",
        }
    )

    return pd.DataFrame(rows)


def write_master_workbook(
    output_xlsx: Path,
    source_df: pd.DataFrame,
    f31_df: pd.DataFrame,
    f33_df: pd.DataFrame,
    a31_df: pd.DataFrame,
    a32_df: pd.DataFrame,
    p31_df: pd.DataFrame,
    p33_df: pd.DataFrame,
    activity_link_df: pd.DataFrame,
    activity_quarter_df: pd.DataFrame,
    project_quarter_df: pd.DataFrame,
    narrative_mentions_df: pd.DataFrame,
    narrative_quarters_df: pd.DataFrame,
    timeline_df: pd.DataFrame,
    coverage_df: pd.DataFrame,
) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        readme_df = pd.DataFrame(
            [
                {
                    "section": "purpose",
                    "detail": "Master workbook with normalized source tables and explicit link audit tabs.",
                },
                {
                    "section": "canonical_sources",
                    "detail": "Includes only one canonical copy per report key; ignores duplicate [n] downloads and ._* sidecars.",
                },
                {
                    "section": "key_1",
                    "detail": "canonical_activity_id",
                },
                {
                    "section": "key_2",
                    "detail": "canonical_activity_id + quarter_label",
                },
                {
                    "section": "key_3",
                    "detail": "grant_number + project_title_norm + quarter_label",
                },
                {
                    "section": "narrative_bridge",
                    "detail": "A31/A32 narrative mentions are used to classify F33-only zero rows as inferred structural zeros.",
                },
                {
                    "section": "limitation",
                    "detail": "P33 is not quarterized in source export, so quarter-level joins to P33 are not possible.",
                },
            ]
        )
        readme_df.to_excel(writer, index=False, sheet_name="README")
        source_df.to_excel(writer, index=False, sheet_name="source_files")
        f31_df.to_excel(writer, index=False, sheet_name="F31_activity_fin")
        f33_df.to_excel(writer, index=False, sheet_name="F33_project_qtr")
        a31_df.to_excel(writer, index=False, sheet_name="A31_exec_narr")
        a32_df.to_excel(writer, index=False, sheet_name="A32_activity_narr")
        p31_df.to_excel(writer, index=False, sheet_name="P31_accomp_qtr")
        p33_df.to_excel(writer, index=False, sheet_name="P33_beneficiary")
        activity_link_df.to_excel(writer, index=False, sheet_name="link_activity")
        activity_quarter_df.to_excel(writer, index=False, sheet_name="link_activity_qtr")
        project_quarter_df.to_excel(writer, index=False, sheet_name="link_project_qtr")
        narrative_mentions_df.to_excel(writer, index=False, sheet_name="narr_project_mentions")
        narrative_quarters_df.to_excel(writer, index=False, sheet_name="narr_quarter_coverage")
        timeline_df.to_excel(writer, index=False, sheet_name="link_quarter_tl")
        coverage_df.to_excel(writer, index=False, sheet_name="link_coverage")

    wb = load_workbook(output_xlsx)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions
    wb.save(output_xlsx)


def build_master(input_root: Path, output_xlsx: Path, output_summary_csv: Path) -> None:
    sources = resolve_source_files(input_root)
    source_rows = []
    for spec in SOURCE_SPECS:
        key = spec["key"]
        source_rows.append(
            {
                "source_key": key,
                "report_type": spec["report"],
                "grant_number": spec["grant"],
                "file_path": str(sources[key]),
                "file_name": sources[key].name,
            }
        )
    source_df = pd.DataFrame(source_rows)

    f31_df = pd.concat(
        [
            parse_f31(sources["B17_F31"], "B17_F31", "B-17-DM-48-0001"),
            parse_f31(sources["B18_F31"], "B18_F31", "B-18-DP-48-0001"),
        ],
        ignore_index=True,
    )
    f33_df = pd.concat(
        [
            parse_f33(sources["B17_F33"], "B17_F33", "B-17-DM-48-0001"),
            parse_f33(sources["B18_F33"], "B18_F33", "B-18-DP-48-0001"),
        ],
        ignore_index=True,
    )
    a31_df = parse_a31(sources["HIM1_A31"], "HIM1_A31", "P-17-TX-48-HIM1")
    a32_df = parse_a32(sources["HIM1_A32"], "HIM1_A32", "P-17-TX-48-HIM1")
    p31_df = parse_p31(sources["HIM1_P31"], "HIM1_P31", "P-17-TX-48-HIM1")
    p33_df = parse_p33(sources["HIM1_P33"], "HIM1_P33", "P-17-TX-48-HIM1")

    activity_link_df = build_activity_link(f31_df, a32_df, p31_df, p33_df)
    activity_quarter_df = build_activity_quarter_link(f31_df, a32_df, p31_df)
    narrative_mentions_df, narrative_quarters_df = build_narrative_project_mentions(
        a31_df, a32_df, f33_df
    )
    project_quarter_df = build_project_quarter_link(
        f31_df, f33_df, narrative_mentions_df, narrative_quarters_df
    )
    timeline_df = build_quarter_timeline_link(f31_df, f33_df, a31_df, a32_df, p31_df)
    coverage_df = build_coverage_summary(
        activity_link_df, activity_quarter_df, project_quarter_df, timeline_df
    )

    write_master_workbook(
        output_xlsx=output_xlsx,
        source_df=source_df,
        f31_df=f31_df,
        f33_df=f33_df,
        a31_df=a31_df,
        a32_df=a32_df,
        p31_df=p31_df,
        p33_df=p33_df,
        activity_link_df=activity_link_df,
        activity_quarter_df=activity_quarter_df,
        project_quarter_df=project_quarter_df,
        narrative_mentions_df=narrative_mentions_df,
        narrative_quarters_df=narrative_quarters_df,
        timeline_df=timeline_df,
        coverage_df=coverage_df,
    )

    output_summary_csv.parent.mkdir(parents=True, exist_ok=True)
    coverage_df.to_csv(output_summary_csv, index=False)

    print(f"Wrote master workbook: {output_xlsx}")
    print(f"Wrote coverage summary CSV: {output_summary_csv}")
    print(
        "Row counts -> "
        f"F31={len(f31_df):,}, F33={len(f33_df):,}, A31={len(a31_df):,}, "
        f"A32={len(a32_df):,}, P31={len(p31_df):,}, P33={len(p33_df):,}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build linked master workbook from QPR XLSX files.")
    parser.add_argument(
        "--input-root",
        type=Path,
        default=ROOT,
        help="Directory that contains the QPR XLSX files.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=ROOT / "output" / "spreadsheet" / "Master_QPR_Linked.xlsx",
        help="Path for the output master workbook.",
    )
    parser.add_argument(
        "--output-summary-csv",
        type=Path,
        default=ROOT / "output" / "spreadsheet" / "Master_QPR_Link_Coverage.csv",
        help="Path for linkage coverage summary CSV.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    build_master(
        input_root=args.input_root,
        output_xlsx=args.output_xlsx,
        output_summary_csv=args.output_summary_csv,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
