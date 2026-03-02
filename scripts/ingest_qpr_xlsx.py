#!/usr/bin/env python3
"""
Ingest structured QPR XLSX exports into the project database.

Reads 8 unique XLSX files (F31, F33, A31, A32, P31, P33) from the project root
and populates four database tables:

  - qpr_activity_financials   (from F31 files)
  - qpr_payroll_allocations   (from A32 activity narratives)
  - qpr_accomplishments       (from P31 file)
  - qpr_beneficiary_demographics (from P33 file)

Usage:
    python scripts/ingest_qpr_xlsx.py              # incremental insert
    python scripts/ingest_qpr_xlsx.py --rebuild    # drop + re-insert
    python scripts/ingest_qpr_xlsx.py --stats      # show table counts
"""

from __future__ import annotations

import argparse
import re
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

import sys as _sys
_sys.path.insert(0, str(ROOT / "src"))
try:
    from utils import parse_usd as _parse_usd
except ImportError:
    _parse_usd = None

# ---------------------------------------------------------------------------
# File map: logical key -> expected filename in ROOT
# ---------------------------------------------------------------------------
XLSX_FILES: Dict[str, str] = {
    "B17_F31": "B-17-DM-48-0001_F31 - QPR - Fin Data by Project, Activity and Quarter_02032026.xlsx",
    "B18_F31": "B-18-DP-48-0001_F31 - QPR - Fin Data by Project, Activity and Quarter_02032026.xlsx",
    "B17_F33": "B-17-DM-48-0001_F33 - QPR - Fin Data by Quarter - Project Level_02032026.xlsx",
    "B18_F33": "B-18-DP-48-0001_F33 - QPR - Fin Data by Quarter - Project Level_02032026.xlsx",
    "HIM1_A31": "P-17-TX-48-HIM1_A31 - QPR -\u00a0 Executive Summary Progress Narrative_02032026.xlsx",
    "HIM1_A32": "P-17-TX-48-HIM1_A32 - QPR - Activity Progress Narratives_02032026.xlsx",
    "HIM1_P31": "P-17-TX-48-HIM1_P31 - QPR - Actual Accomplishments by Quarter_02032026.xlsx",
    "HIM1_P33": "P-17-TX-48-HIM1_P33 - QPR - Household Characteristics for Direct Benefit Activities by Tenure and Ethnicity_02032026.xlsx",
}

GRANT_FOR_KEY: Dict[str, str] = {
    "B17_F31": "B-17-DM-48-0001",
    "B18_F31": "B-18-DP-48-0001",
    "B17_F33": "B-17-DM-48-0001",
    "B18_F33": "B-18-DP-48-0001",
    "HIM1_A31": "P-17-TX-48-HIM1",
    "HIM1_A32": "P-17-TX-48-HIM1",
    "HIM1_P31": "P-17-TX-48-HIM1",
    "HIM1_P33": "P-17-TX-48-HIM1",
}

QPR_TABLES = [
    "qpr_activity_financials",
    "qpr_payroll_allocations",
    "qpr_accomplishments",
    "qpr_beneficiary_demographics",
]

# Regex for extracting payroll amounts from A32 narratives
_PAYROLL_AMOUNT_RE = re.compile(
    r"(?:^|\s)(\$?\s*\d[\d,]*(?:\.\d+)?)\s*[-\u2013\u2014]\s*[Pp]ayroll",
)
_MONEY_IN_PAYROLL_LINE_RE = re.compile(
    r"(?:US\$|\$)\s*\d[\d,]*(?:\.\d+)?|\d[\d,]*(?:\.\d+)?\s*(?:million|billion)",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _resolve_file(input_root: Path, expected_name: str) -> Optional[Path]:
    """Find XLSX file, tolerating bracket-suffixed download duplicates."""
    direct = input_root / expected_name
    if direct.exists():
        return direct
    base = expected_name.rsplit(".", 1)[0]
    pattern = f"{base}[[]*.xlsx"
    candidates = sorted(input_root.glob(pattern))
    if candidates:
        return candidates[0]
    return None


def _open_sheet(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    return wb.worksheets[0]


def _nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _clean_str(value: Any) -> Optional[str]:
    if not _nonempty(value):
        return None
    return str(value).replace("\xa0", " ").strip()


def _clean_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value == value else None  # NaN check
    # Delegate string parsing to project-wide helper (CLAUDE.md convention)
    if _parse_usd is not None:
        return _parse_usd(str(value).strip())
    # Fallback (should not occur in normal use):
    s = str(value).strip().replace(",", "").replace("$", "")
    return float(s) if s else None


def _quarter_label(value: Any) -> Optional[str]:
    """Convert a datetime or string quarter reference to 'YYYY QN' label."""
    if isinstance(value, datetime):
        q = (value.month - 1) // 3 + 1
        return f"{value.year} Q{q}"
    s = _clean_str(value)
    if s and re.match(r"\d{4}\s*Q\d", s):
        return s.strip()
    return s


def _date_str(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _clean_str(value)


def _parse_payroll_usd(text: str) -> List[float]:
    """Extract payroll dollar amounts from a narrative line."""
    amounts: List[float] = []
    # Pattern 1: "$89382.76 - Payroll" or "89382.76 - Payroll"
    for m in _PAYROLL_AMOUNT_RE.finditer(text):
        raw = m.group(1).strip().replace(",", "").replace("$", "").strip()
        try:
            v = float(raw)
            if v > 0:
                amounts.append(v)
        except ValueError:
            pass
    # Pattern 2: broader money tokens in payroll-containing lines
    if not amounts and "payroll" in text.lower():
        for m in _MONEY_IN_PAYROLL_LINE_RE.finditer(text):
            raw = m.group(0).strip().replace(",", "").replace("$", "").strip()
            # Handle magnitude suffixes
            low = raw.lower()
            mult = 1.0
            if low.endswith("million"):
                raw = raw[: -len("million")].strip()
                mult = 1_000_000
            elif low.endswith("billion"):
                raw = raw[: -len("billion")].strip()
                mult = 1_000_000_000
            try:
                v = float(raw) * mult
                if v > 0:
                    amounts.append(v)
            except ValueError:
                pass
    return amounts


# ---------------------------------------------------------------------------
# Ingestion functions
# ---------------------------------------------------------------------------

def ingest_f31(conn: sqlite3.Connection, input_root: Path, file_key: str) -> int:
    """Ingest F31 (Financial Data by Project, Activity and Quarter) file."""
    grant = GRANT_FOR_KEY[file_key]
    path = _resolve_file(input_root, XLSX_FILES[file_key])
    if path is None:
        print(f"  SKIP {file_key}: file not found")
        return 0

    ws = _open_sheet(path)
    source_file = path.name
    rows_inserted = 0
    cursor = conn.cursor()

    for row in ws.iter_rows(min_row=11, max_col=13, values_only=True):
        activity_number = _clean_str(row[2])
        if not activity_number:
            continue
        # Skip subtotal/grand total rows
        project_number = _clean_str(row[0])
        if not project_number and not activity_number:
            continue

        quarter = _quarter_label(row[7]) or _quarter_label(row[6])
        if not quarter:
            continue

        try:
            cursor.execute(
                """INSERT OR REPLACE INTO qpr_activity_financials
                   (grant_number, project_number, project_title, activity_number,
                    activity_title, activity_type, responsible_org, begin_date,
                    quarter_label, obligated_usd, expended_usd, disbursed_usd,
                    program_income_disbursed_usd, program_income_received_usd,
                    source_file)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    grant,
                    project_number,
                    _clean_str(row[1]),
                    activity_number,
                    _clean_str(row[3]),
                    _clean_str(row[4]),
                    _clean_str(row[5]),
                    _date_str(row[6]),
                    quarter,
                    _clean_float(row[8]),
                    _clean_float(row[9]),
                    _clean_float(row[10]),
                    _clean_float(row[11]),
                    _clean_float(row[12]),
                    source_file,
                ),
            )
            rows_inserted += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    return rows_inserted


def ingest_a32_payroll(conn: sqlite3.Connection, input_root: Path) -> int:
    """Extract payroll allocations from A32 (Activity Progress Narratives)."""
    file_key = "HIM1_A32"
    grant = GRANT_FOR_KEY[file_key]
    path = _resolve_file(input_root, XLSX_FILES[file_key])
    if path is None:
        print(f"  SKIP {file_key}: file not found")
        return 0

    ws = _open_sheet(path)
    source_file = path.name
    rows_inserted = 0
    cursor = conn.cursor()

    for row in ws.iter_rows(min_row=9, max_col=7, values_only=True):
        activity_number = _clean_str(row[0])
        if not activity_number:
            continue
        narrative = _clean_str(row[6])
        if not narrative:
            continue
        if "payroll" not in narrative.lower():
            continue

        quarter = _quarter_label(row[5]) or _quarter_label(row[4])
        if not quarter:
            continue

        amounts = _parse_payroll_usd(narrative)
        for amount in amounts:
            try:
                cursor.execute(
                    """INSERT OR REPLACE INTO qpr_payroll_allocations
                       (grant_number, activity_number, activity_title,
                        activity_type, responsible_org, quarter_label,
                        payroll_usd, narrative_snippet, source_file)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (
                        grant,
                        activity_number,
                        _clean_str(row[1]),
                        _clean_str(row[2]),
                        _clean_str(row[3]),
                        quarter,
                        amount,
                        narrative[:500],
                        source_file,
                    ),
                )
                rows_inserted += 1
            except sqlite3.IntegrityError:
                pass

    conn.commit()
    return rows_inserted


def ingest_p31(conn: sqlite3.Connection, input_root: Path) -> int:
    """Ingest P31 (Actual Accomplishments by Quarter) file."""
    file_key = "HIM1_P31"
    grant = GRANT_FOR_KEY[file_key]
    path = _resolve_file(input_root, XLSX_FILES[file_key])
    if path is None:
        print(f"  SKIP {file_key}: file not found")
        return 0

    ws = _open_sheet(path)
    source_file = path.name
    cursor = conn.cursor()
    rows_inserted = 0

    # Row 8 has quarter header dates (datetime objects)
    header_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    # Row 9 has quarter labels like "2020 Q2"
    label_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    # Build quarter label map for columns 6+ (0-indexed)
    quarter_cols: List[Tuple[int, str]] = []
    for col_idx in range(6, len(label_row)):
        qlabel = _quarter_label(label_row[col_idx]) or _quarter_label(header_row[col_idx])
        if qlabel:
            quarter_cols.append((col_idx, qlabel))

    for row in ws.iter_rows(min_row=10, values_only=True):
        activity_number = _clean_str(row[2])
        if not activity_number:
            continue
        measure_type = _clean_str(row[4])
        if not measure_type:
            continue

        for col_idx, qlabel in quarter_cols:
            if col_idx >= len(row):
                continue
            val = _clean_float(row[col_idx])
            if val is None:
                continue
            try:
                cursor.execute(
                    """INSERT OR REPLACE INTO qpr_accomplishments
                       (grant_number, activity_number, activity_title,
                        activity_type, responsible_org, measure_type,
                        quarter_label, actual_value, source_file)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (
                        grant,
                        activity_number,
                        _clean_str(row[3]),
                        _clean_str(row[0]),
                        _clean_str(row[1]),
                        measure_type,
                        qlabel,
                        val,
                        source_file,
                    ),
                )
                rows_inserted += 1
            except sqlite3.IntegrityError:
                pass

    conn.commit()
    return rows_inserted


def ingest_p33(conn: sqlite3.Connection, input_root: Path) -> int:
    """Ingest P33 (Household Characteristics by Tenure and Ethnicity)."""
    file_key = "HIM1_P33"
    grant = GRANT_FOR_KEY[file_key]
    path = _resolve_file(input_root, XLSX_FILES[file_key])
    if path is None:
        print(f"  SKIP {file_key}: file not found")
        return 0

    ws = _open_sheet(path)
    source_file = path.name
    cursor = conn.cursor()
    rows_inserted = 0

    # Row 8 has category headers (e.g., "Households - Owner Total")
    # Row 9 has demographic sub-group headers (e.g., "White", "Black/African American")
    cat_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    demo_row = list(ws.iter_rows(min_row=9, max_row=9, values_only=True))[0]

    # Build column map: col_idx -> (category, demographic_group)
    col_map: List[Tuple[int, str, str]] = []
    current_category = ""
    for col_idx in range(5, min(len(cat_row), len(demo_row))):
        cat = _clean_str(cat_row[col_idx])
        if cat:
            current_category = cat
        demo = _clean_str(demo_row[col_idx])
        if current_category and demo:
            col_map.append((col_idx, current_category, demo))

    for row in ws.iter_rows(min_row=11, values_only=True):
        activity_number = _clean_str(row[1])
        if not activity_number:
            continue

        for col_idx, category, demographic_group in col_map:
            if col_idx >= len(row):
                continue
            val = _clean_float(row[col_idx])
            if val is None:
                continue
            try:
                cursor.execute(
                    """INSERT OR REPLACE INTO qpr_beneficiary_demographics
                       (grant_number, activity_number, activity_title,
                        activity_type, national_objective, category,
                        demographic_group, value, source_file)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (
                        grant,
                        activity_number,
                        _clean_str(row[2]),
                        _clean_str(row[0]),
                        _clean_str(row[3]),
                        category,
                        demographic_group,
                        val,
                        source_file,
                    ),
                )
                rows_inserted += 1
            except sqlite3.IntegrityError:
                pass

    conn.commit()
    return rows_inserted


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def _get_db_path() -> Path:
    try:
        import sys
        sys.path.insert(0, str(ROOT / "src"))
        import config
        return config.DATABASE_PATH
    except Exception:
        return ROOT / "data" / "glo_reports.db"


def _init_db(db_path: Path) -> sqlite3.Connection:
    import sys
    sys.path.insert(0, str(ROOT / "src"))
    from utils import init_database
    return init_database(db_path)


def _table_count(conn: sqlite3.Connection, table: str) -> int:
    try:
        return conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    except sqlite3.OperationalError:
        return 0


def show_stats(conn: sqlite3.Connection) -> None:
    print("\nQPR XLSX Table Counts:")
    for table in QPR_TABLES:
        print(f"  {table}: {_table_count(conn, table):,}")


def rebuild_tables(conn: sqlite3.Connection) -> None:
    cursor = conn.cursor()
    for table in QPR_TABLES:
        cursor.execute(f"DELETE FROM {table}")
        print(f"  Cleared {table}")
    conn.commit()


def run_ingestion(input_root: Path, conn: sqlite3.Connection) -> Dict[str, int]:
    counts: Dict[str, int] = {}

    print("\n1. Ingesting F31 financial data...")
    n = ingest_f31(conn, input_root, "B17_F31")
    print(f"   B-17 F31: {n:,} rows")
    counts["B17_F31"] = n

    n = ingest_f31(conn, input_root, "B18_F31")
    print(f"   B-18 F31: {n:,} rows")
    counts["B18_F31"] = n

    print("\n2. Extracting payroll allocations from A32 narratives...")
    n = ingest_a32_payroll(conn, input_root)
    print(f"   A32 payroll: {n:,} rows")
    counts["A32_payroll"] = n

    print("\n3. Ingesting P31 accomplishment data...")
    n = ingest_p31(conn, input_root)
    print(f"   P31 accomplishments: {n:,} rows")
    counts["P31"] = n

    print("\n4. Ingesting P33 beneficiary demographics...")
    n = ingest_p33(conn, input_root)
    print(f"   P33 demographics: {n:,} rows")
    counts["P33"] = n

    return counts


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Ingest QPR XLSX exports into the project database."
    )
    parser.add_argument(
        "--input-root",
        type=Path,
        default=ROOT,
        help="Directory containing the QPR XLSX exports.",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=None,
        help="Database path (defaults to config.DATABASE_PATH).",
    )
    parser.add_argument(
        "--rebuild",
        action="store_true",
        help="Clear all QPR tables before ingesting.",
    )
    parser.add_argument(
        "--stats",
        action="store_true",
        help="Show table counts and exit.",
    )
    args = parser.parse_args(argv)

    db_path = args.db or _get_db_path()
    conn = _init_db(db_path)

    if args.stats:
        show_stats(conn)
        conn.close()
        return 0

    print("=" * 60)
    print("INGESTING QPR XLSX EXPORTS")
    print("=" * 60)

    if args.rebuild:
        print("\nClearing existing QPR tables...")
        rebuild_tables(conn)

    counts = run_ingestion(args.input_root, conn)

    print("\n" + "=" * 60)
    print("INGESTION COMPLETE")
    print("=" * 60)
    show_stats(conn)
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
